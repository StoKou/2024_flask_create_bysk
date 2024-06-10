from flask import Flask, jsonify, request, make_response
from flask_cors import CORS,cross_origin
from openpyxl import load_workbook
from app.Model.function.model1 import process_model_1
from app.Model.function.model2 import process_model_2
from app.Model.function.data_images_get_1 import data_images_get_1
from app.Model.function.get_path import get_path

app = Flask(__name__)
CORS(app, resources={r'/*': {'origins': '*'}})  # 允许所有域名跨域访问，可根据需求细化控制

# 假设这里是您从数据库或其他来源获取的数据
sample_home_data = { 
    "title": "首页标题",
    "description": "<p>欢迎来到我们的网站！这里有一些关于我们的pi介绍和最新动态。</p>"
}

@app.route('/api/predict/model_1', methods=['POST'])
def predict_model_1():
    file = request.files['file']
    result = process_model_1(file)
    return jsonify(result), 200

@app.route('/api/predict/model_2', methods=['POST'])
def predict_model_2():
    file = request.files['file']
    result = process_model_2(file)
    return jsonify(result), 200
@app.after_request
def add_cors_headers(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
    return response

@app.route('/api/home', methods=['GET'])
@cross_origin()
def get_home_data():
    return jsonify(sample_home_data)

@app.route('/api/expimage', methods=['GET'])
def get_excel_data():
    # 加载 Excel 文件
    workbook = load_workbook(filename='./test.xlsx')
    sheet = workbook.active  # 获取第一个工作表

    # 将 Excel 数据转换为列表
    categories = [cell.value for cell in sheet['A'] if cell.value is not None]
    values = [cell.value for row in sheet.iter_rows(min_row=2, max_col=2) for cell in row if cell.value is not None]

    # 转换为 JSON 格式并返回
    data = {
        'categories': categories,
        'values': values
    }

    # 设置允许跨域的响应头信息（虽然在使用了 CORS 插件后通常不需要手动设置）
    response = jsonify(data)
    # response.headers.add('Access-Control-Allow-Origin', '*')  # 允许所有域名访问，也可指定特定域名
    # response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')  # 允许自定义请求头
    # response.headers.add('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')  # 允许的方法

    return response

@app.route('/api/expimage/get_1')
def get_image_data():
    # 指定.mat文件的路径
    img_path = get_path(option="images")
    
     # 调用data_images_get_1函数获取数据
    data = data_images_get_1(img_path)
    
    # 根据data字典中是否存在'error'键来判断是否发生了错误
    if 'error' in data:
        # 发生了错误，返回错误信息和500状态码
        return jsonify(data), 500
    else:
        # 没有错误，正常返回数据
        return jsonify(data)

if __name__ == '__main__':
    app.run(debug=True)